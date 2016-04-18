from Tkinter import Tk, Text,X, LEFT
from ttk import Frame, Label, Entry, Button, Radiobutton
from Tkinter import BOTH,StringVar
import tkFileDialog
import os
import openpyxl
from openpyxl.styles import Font, Alignment



class Example(Frame):
  
    def __init__(self, parent):
        Frame.__init__(self, parent)   
         
        self.parent = parent
        
        self.initUI()
        
        
    def initUI(self):
        self.parent.title("Filter Data")
        self.pack(fill=BOTH, expand=True)
        labelfont20 = ('Roboto', 15, 'bold')
        labelfont10 = ('Roboto', 10, 'bold')
        labelfont8 = ('Roboto', 8, 'bold')
        
        frame0 = Frame(self)
        frame0.pack()
        
        lbl0 = Label(frame0, text="Hi Nakul")
        lbl0.config(font=labelfont20)    
        lbl0.pack( padx=5, pady=5)
        lbl00 = Label(frame0, text="Filter Data")
        lbl00.config(font=labelfont10)
        lbl00.pack( padx=5, pady=5)
        
        ####################################
        
        
        ##########################################
        
        
        ############################################
        #####printing line
        
        lbl5a = Label(text="__________________________________")
        lbl5a.pack()
        lbl5a.place(x=170, y=300)
        
        lbl5b = Label(text="__________________________________")
        lbl5b.pack()
        lbl5b.place(x=480, y=300)
        
        self.lbl5c = Label(text="Search Result Will Appear Here")
        self.lbl5c.pack()
        self.lbl5c.place(x=170, y=320)
        
        self.lbl5d = Label(text="File Name Will Appear Here")
        self.lbl5d.pack()
        self.lbl5d.place(x=480, y=320)
        
        ############################################
        
        
        #############################################
        
        ###############################################
        
        
        ##############################################
        
        #############################################
        
        frame11 = Frame(self)
        frame11.pack()
        frame11.place(x=200, y=100)
        
        lbl11x = Label(frame11,text="Class 10th %")
        lbl11x.pack(padx=0, pady=0)
        
               

        self.entry11 = Entry(width=12)
        self.entry11.pack(padx=1, expand=True)
        self.entry11.place(x=200, y=120) 
        
        
        
        
        ####################################################
        frame12 = Frame(self)
        frame12.pack()
        frame12.place(x=380, y=100)
        
        lbl12x = Label(frame12,text="Class 12th %")
        lbl12x.pack(padx=0, pady=0)
        
               

        self.entry12 = Entry(width=12)
        self.entry12.pack(padx=1, expand=True)
        self.entry12.place(x=380, y=120) 
        
        

        #####################################################
        frame13 = Frame(self)
        frame13.pack()
        frame13.place(x=550, y=100)
        
        lbl13x = Label(frame13,text="B.Tech %")
        lbl13x.pack(padx=0, pady=0)
        
               

        self.entry13 = Entry(width=12)
        self.entry13.pack(padx=1, expand=True)
        self.entry13.place(x=550, y=120) 
        
        
        
        ####################################################
        frame9 = Frame(self)
        frame9.pack()
        frame9.place(x=350, y=160)
        
        lbl9 = Label(frame9, text="HomeTown:")
        lbl9.pack()        

        self.entry9 = Entry(frame9)
        self.entry9.pack(fill=X, padx=5, expand=True)
        
        
             
         
        
        
        #############################################################
        frame16 = Frame(self)
        frame16.pack()
        frame16.place(x=190, y=250)
        closeButton = Button(frame16, text="Filter",width=20,command=self.getDatax2)
        closeButton.pack(padx=5, pady=5)
        
        #######################################
        frame17 = Frame(self)
        frame17.pack()
        frame17.place(x=500, y=250)
        closeButton = Button(frame17, text="Save & Open",width=20,command=self.getDatax3)
        closeButton.pack(padx=5, pady=5)
        
        #######################################
        
        frame000 = Frame(self)
        frame000.pack()
        frame000.place(x=50, y=600)
        
        self.lbl000= Label(frame000, text="Beta/Sample2.0 | (c) Nakul Rathore")
        self.lbl000.config(font=labelfont8)    
        self.lbl000.pack( padx=5, pady=5)
        
        
        
    def getDatax2(self):
        x1 = self.entry11.get()
        if x1 != "":
            x1 = int(x1)
        
        x2 = self.entry12.get()
        if x2 != "":
            x2 = int(x2)
        x3 = self.entry13.get()
        if x3 != "":
            x3 = int(x3)
        x4 = self.entry9.get()
        list1=[x1,x2,x3,x4]
        
        wb = openpyxl.load_workbook('..\database\database.xlsx')
        ws = wb.active
        print(wb.get_sheet_names())
        max_row = ws.get_highest_row()
        max_col = ws.get_highest_column()
        global temp
        global tempx
        temp = []
        tempx = []
        for i in xrange(2,max_row+1):
            temp.append(i)
        #print temp
        
        if isinstance(x1, int):
            for i in temp:
                if ws.cell(row = i, column = 11).value >= x1:
                    tempx.append(i)
            temp = tempx
            tempx = []
            print temp
            
        if isinstance(x2, int):
            for i in temp:
                if ws.cell(row = i, column = 14).value >= x2:
                    tempx.append(i)
            temp = tempx
            tempx = []
            print temp
        if isinstance(x3, int):
            for i in temp:
                if ws.cell(row = i, column = 17).value >= x3:
                    tempx.append(i)
            temp = tempx
            tempx = []
            print temp
            
        if isinstance(x3, str) and x3 != "":
            for i in temp:
                if ws.cell(row = i, column = 9).value == x4:
                    tempx.append(i)
            temp = tempx
            tempx = []
            print temp
        self.lbl5c.config(text=""+str(len(temp))+" result(s) found")
            
    def getDatax3(self):
        import datetime
        now = datetime.datetime.now()
        now = now.replace(microsecond=0,second = 0)
        now = now.strftime("%d_%B_%y,%I-%M_%p")
        now = now+".xlsx"
        
       
        
        
        if len(temp) != 0:
            wb1 = openpyxl.load_workbook('..\database\database.xlsx')
            ws1 = wb1.active
            wb2 = openpyxl.load_workbook('..\_frame\_frame.xlsx')
            ws2 = wb2.active
        
            for i in xrange(2,len(temp)+2):
                for j in xrange(1,22):
                    ws2.cell(row = i, column = j).value = ws1.cell(row = temp[i-2], column = j).value
        
        wb2.save('..\Result\\'+now)
        tempstart = '..\Result\\'+now
        self.lbl5d.config(text="File is :: "+"\""+now+"\"")
        os.system("start "+tempstart)
        
        self.entry11.delete(0, 'end')
        self.entry12.delete(0, 'end')
        self.entry13.delete(0, 'end')
        self.entry9.delete(0, 'end')
        
        
        
                
                
        
        
    

        


def main():
  
    root = Tk()
    
    root.geometry("900x650+300+20")
    app = Example(root)
    root.mainloop()  


if __name__ == '__main__':
    main() 