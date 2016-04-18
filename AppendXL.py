from Tkinter import Tk, Text,X, LEFT
from ttk import Frame, Label, Entry, Button, Radiobutton
from Tkinter import BOTH,StringVar
import openpyxl
import tkFileDialog
import os
from openpyxl.styles import Font, Alignment


class Example(Frame):
  
    def __init__(self, parent):
        Frame.__init__(self, parent)   
         
        self.parent = parent
        
        self.initUI()
        
        
    def initUI(self):
   
      
        self.parent.title("Append Data")
        self.pack(fill=BOTH, expand=True)
        labelfont20 = ('Roboto', 15, 'bold')
        labelfont10 = ('Roboto', 10, 'bold')
        labelfont8 = ('Roboto', 8, 'bold')
        
        frame0 = Frame(self)
        frame0.pack()
        
        lbl0 = Label(frame0, text="Hi Nakul")
        lbl0.config(font=labelfont20)    
        lbl0.pack( padx=5, pady=5)
        lbl00 = Label(frame0, text="Fill the data here")
        lbl00.config(font=labelfont10)
        lbl00.pack( padx=5, pady=5)
        
        ####################################
        frame1 = Frame(self)
        frame1.pack()
        frame1.place(x=50, y=100)        
        
        lbl1 = Label(frame1, text="Name", width=15)
        lbl1.pack(side=LEFT,padx=7, pady=5) 
             
        self.entry1 = Entry(frame1,width=20)
        self.entry1.pack(padx=5, expand=True)
    
        ####################################
        frame2 = Frame(self)
        frame2.pack()
        frame2.place(x=50, y=130)
        
        lbl2 = Label(frame2, text="F Name", width=15)
        lbl2.pack(side=LEFT, padx=7, pady=5)

        self.entry2 = Entry(frame2)
        self.entry2.pack(fill=X, padx=5, expand=True)
        
        ######################################
        frame3 = Frame(self)
        frame3.pack()
        frame3.place(x=50, y=160)
        
        lbl3 = Label(frame3, text="DOB(D/M/Y)", width=15)
        lbl3.pack(side=LEFT, padx=7, pady=5)        

        self.entry3 = Entry(frame3)
        self.entry3.pack(fill=X, padx=5, expand=True) 
        
        #######################################
        frame4 = Frame(self)
        frame4.pack()
        frame4.place(x=50, y=190)
        
        lbl4 = Label(frame4, text="Medium(H/E)", width=15)
        lbl4.pack(side=LEFT, padx=7, pady=5)        

        self.entry4 = Entry(frame4)
        self.entry4.pack(fill=X, padx=5, expand=True)
        
        ##########################################
        frame5 = Frame(self)
        frame5.pack()
        frame5.place(x=50, y=225)  
        MODES = [
            ("M", "Male"),
            ("F", "Female"),
            ]
        lbl5 = Label(frame5, text="Gender", width=15)
        lbl5.pack(side=LEFT, padx=7, pady=5)

        global v
        v = StringVar()
        v.set("Male") # initialize

        for text, mode in MODES:
            b = Radiobutton(frame5, text=text,variable=v, value=mode)
            b.pack(side=LEFT,padx=10)
        
        ############################################
        #####printing line
        lbl5a = Label(text="___________________________________________________")
        lbl5a.pack()
        lbl5a.place(x=45, y=255)  
        
        ############################################
        frame6 = Frame(self)
        frame6.pack()
        frame6.place(x=50, y=290)
        
        lbl6 = Label(frame6, text="Phone No:", width=15)
        lbl6.pack(side=LEFT, padx=7, pady=5)        

        self.entry6 = Entry(frame6)
        self.entry6.pack(fill=X, padx=5, expand=True)
        
        ################################################
        
        frame7 = Frame(self)
        frame7.pack()
        frame7.place(x=50, y=320)
        
        lbl7 = Label(frame7, text="Landline No:", width=15)
        lbl7.pack(side=LEFT, padx=7, pady=5)        

        self.entry7 = Entry(frame7)
        self.entry7.pack(fill=X, padx=5, expand=True)
        
        ###############################################
        frame8 = Frame(self)
        frame8.pack()
        frame8.place(x=50, y=350)
        
        lbl8 = Label(frame8, text="Email:", width=15)
        lbl8.pack(side=LEFT, padx=7, pady=5)        

        self.entry8 = Entry(frame8)
        self.entry8.pack(fill=X, padx=5, expand=True)
        
        #############################################
        frame9 = Frame(self)
        frame9.pack()
        frame9.place(x=50, y=380)
        
        lbl9 = Label(frame9, text="HomeTown:", width=15)
        lbl9.pack(side=LEFT, padx=7, pady=5)        

        self.entry9 = Entry(frame9)
        self.entry9.pack(fill=X, padx=5, expand=True)
        
        ###############################################
        frame10 = Frame(self)
        frame10.pack()
        frame10.place(x=60, y=415)
        
        lbl10 = Label(frame10, text="Address:")
        lbl10.pack( padx=5, pady=5)        

        self.entry10 = Text(frame10,height=5, width=28)
        self.entry10.pack(padx=5, expand=True)
        
        ##############################################
        
        #############################################
        
        frame11 = Frame(self)
        frame11.pack()
        frame11.place(x=350, y=100)
        
        lbl11x = Label(frame11,text="_______Class 10th Data_______")
        lbl11x.pack(padx=0, pady=0)
        
        lbl11 = Label(text="%",width=15)
        lbl11.pack(side=LEFT,padx=0, pady=0)
        lbl11.place(x=350, y=130)        

        self.entry11 = Entry(width=12)
        self.entry11.pack(padx=1, expand=True)
        self.entry11.place(x=420, y=130) 
        
        lbl11a = Label(text="Passing Year",width=15)
        lbl11a.pack(padx=0, pady=2)   
        lbl11a.place(x=350, y=160)   

        self.entry11a = Entry(width=12)
        self.entry11a.pack(padx=1, expand=True)
        self.entry11a.place(x=420, y=160) 
        
        lbl11b = Label(text="Board Name",width=15)
        lbl11b.pack(padx=0, pady=2)   
        lbl11b.place(x=350, y=190)   

        self.entry11b = Entry(width=12)
        self.entry11b.pack(padx=1, expand=True)
        self.entry11b.place(x=420, y=190)
        
        
        ####################################################
        frame12 = Frame(self)
        frame12.pack()
        frame12.place(x=510, y=100)
        
        lbl12x = Label(frame12,text="_______Class 12th Data_______")
        lbl12x.pack(padx=0, pady=0)
        
        lbl12 = Label(text="%",width=15)
        lbl12.pack(side=LEFT,padx=0, pady=0)
        lbl12.place(x=510, y=130)        

        self.entry12 = Entry(width=12)
        self.entry12.pack(padx=1, expand=True)
        self.entry12.place(x=580, y=130) 
        
        lbl12a = Label(text="Passing Year",width=15)
        lbl12a.pack(padx=0, pady=2)   
        lbl12a.place(x=510, y=160)   

        self.entry12a = Entry(width=12)
        self.entry12a.pack(padx=1, expand=True)
        self.entry12a.place(x=580, y=160) 
        
        lbl12b = Label(text="Board Name",width=15)
        lbl12b.pack(padx=0, pady=2)   
        lbl12b.place(x=510, y=190)   

        self.entry12b = Entry(width=12)
        self.entry12b.pack(padx=1, expand=True)
        self.entry12b.place(x=580, y=190)

        #####################################################
        frame13 = Frame(self)
        frame13.pack()
        frame13.place(x=670, y=100)
        
        lbl13x = Label(frame13,text="________B.Tech Data_________")
        lbl13x.pack(padx=0, pady=0)
        
        lbl13 = Label(text="%",width=15)
        lbl13.pack(side=LEFT,padx=0, pady=0)
        lbl13.place(x=670, y=130)        

        self.entry13 = Entry(width=12)
        self.entry13.pack(padx=1, expand=True)
        self.entry13.place(x=740, y=130) 
        
        lbl13a = Label(text="Passing Year",width=15)
        lbl13a.pack(padx=0, pady=2)   
        lbl13a.place(x=670, y=160)   

        self.entry13a = Entry(width=12)
        self.entry13a.pack(padx=1, expand=True)
        self.entry13a.place(x=740, y=160) 
        
        lbl13b = Label(text="College",width=15)
        lbl13b.pack(padx=0, pady=2)   
        lbl13b.place(x=670, y=190)   

        self.entry13b = Entry(width=12)
        self.entry13b.pack(padx=1, expand=True)
        self.entry13b.place(x=740, y=190)
        
        ####################################################
        
        frame14 = Frame(self)
        frame14.pack()
        frame14.place(x=380, y=255)
        
        lbl14 = Label(frame14, text="Any Other Info:")
        lbl14.pack( padx=5, pady=5)        

        self.entry14 = Text(frame14,height=5, width=28)
        self.entry14.pack(padx=5, expand=True)
             
         
        
        frame15 = Frame(self)
        frame15.pack()
        frame15.place(x=650, y=290)
        
        openButton = Button(frame15, text="Attatch Resume",width=15,command=self.openResume)
        openButton.pack(padx=5, pady=5)
        self.entry15 = Entry(frame15)
        self.entry15.pack(fill=X, padx=4, expand=True)
        #############################################################
        frame16 = Frame(self)
        frame16.pack()
        frame16.place(x=450, y=500)
        
        closeButton = Button(frame16, text="SUBMIT",width=35,command=self.getDatax)
        closeButton.pack(padx=5, pady=5)
        
        #######################################
        framexxx = Frame(self)
        framexxx.pack()
        framexxx.place(x=700, y=600)
        self.xxx = Label(framexxx,text="Recent Changes Will Appear Here")
        self.xxx.config(font=labelfont8) 
        self.xxx.pack()
        
        #######################################
        
        frame000 = Frame(self)
        frame000.pack()
        frame000.place(x=50, y=600)
        
        self.lbl000= Label(frame000, text="Beta/Sample2.0 | (c) Nakul Rathore")
        self.lbl000.config(font=labelfont8)    
        self.lbl000.pack( padx=5, pady=5)
        
        
           

    def openResume(self):
        ftypes = [('All files', '*')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes,initialdir='C:/Users/')
        global x15
        fl = dlg.show()
        #file name
        x15 = fl
        temp1 = os.path.basename(fl)
        global temp2
        temp2 = os.path.splitext(temp1)[0]
        
        self.entry15.delete(0, 'end')
        self.entry15.insert(0,temp2)
        
      
        
        
        
        #####################
        
        
        
        
        
    def getDatax(self):
        x1 = self.entry1.get()
        x2 = self.entry2.get()
        x3 = self.entry3.get()
        x4 = self.entry4.get()
        
        x5 = v.get()
        
        x6 = int(self.entry6.get())
        x7 = int(self.entry7.get())
        x8 = self.entry8.get()
        x9 = self.entry9.get()
        
        x10 = self.entry10.get('1.0', 'end')
        
        x11 = int(self.entry11.get())
        x11a = int(self.entry11a.get())
        x11b = self.entry11b.get()
        
        x12 = int(self.entry12.get())
        x12a = int(self.entry12a.get())
        x12b = self.entry12b.get()
        
        x13 = int(self.entry13.get())
        x13a = int(self.entry13a.get())
        x13b = self.entry13b.get()
        
        x14 = self.entry14.get('1.0', 'end')
        
        
        
        
        
        list1=[x1,x2,x3,x4,x5,x6,x7,x8,x9,x10,x11,x11a,x11b,x12,x12a,x12b,x13,x13a,x13b,x14,"=HYPERLINK("+"\""+x15+"\""+","+"\""+temp2+"\""+")"]
        
        
        wb = openpyxl.load_workbook('..\database\database.xlsx')
        ws = wb.active
        print(wb.get_sheet_names())
        max_row = ws.get_highest_row()
        #max_col = ws.get_highest_column()
        max_col = 21
        print max_row
        
        for i in xrange(1,max_col+1):
            #print list1[i]
            ws.cell(row = max_row+1, column = i).value = list1[i-1]
        ws.cell(row = max_row+1, column = max_col).font = Font(color="0000FF", underline='single')
        ws.cell(row = max_row+1, column = max_col).alignment = Alignment(horizontal='center')
        wb.save('..\database\database.xlsx')
        
        
        self.entry1.delete(0, 'end')
        self.entry2.delete(0, 'end')
        self.entry3.delete(0, 'end')
        self.entry4.delete(0, 'end')
        
        self.entry6.delete(0, 'end')
        self.entry7.delete(0, 'end')
        self.entry8.delete(0, 'end')
        self.entry9.delete(0, 'end')
        self.entry10.delete('1.0', '2.0')
        self.entry11.delete(0, 'end')
        self.entry11a.delete(0, 'end')
        self.entry11b.delete(0, 'end')
        self.entry12.delete(0, 'end')
        self.entry12a.delete(0, 'end')
        self.entry12b.delete(0, 'end')
        self.entry13.delete(0, 'end')
        self.entry13a.delete(0, 'end')
        self.entry13b.delete(0, 'end')
        
        self.entry14.delete('1.0', '2.0')
        
        
        self.xxx.config(text="Recent Changes Made For : "+x1)
        
        
    
    
    
    
        


def main():
  
    root = Tk()
    
    
    root.geometry("900x650+300+20")
    app = Example(root)
    root.mainloop()  


if __name__ == '__main__':
    main() 